import os
import sys
import re
import shutil
import codecs
import openpyxl

VERBOSE = False
FEATURE_FILENAME = "{}.feature"
FEATURE_TEMPLATE = """Feature: {}
  {}

  Scenario: {} 
    {}
    {}
    {}
"""

GIVEN_TEMPLATE  = "Given {}"
WHEN_TEMPLATE   = "When  {}"
THEN_TEMPLATE   = "Then  {}"
FEATURE_INDENT  = " " * 2
FILE_NEWLINE    = "\n"
NA              = "NA"


class OpenpyxlReader(object):
    def __init__(self, filename):
        self.__filename = filename

    def __getitem__(self, coordinate):
        return self.get_cell_value(coordinate)
            
    def load(self):
        self.__wb = openpyxl.load_workbook(self.__filename, read_only = True)
        self.__ws = self.__wb.active
        
    def activate_sheet(self, sheetname):
        self.__ws = self.__wb[sheetname]
    
    def get_cell_value(self, coordinate):
        cell = self.__ws[coordinate]
        return [x.value for x in cell] if type(cell) is tuple else cell.value
    
    def get_active_sheet(self):
        return self.__ws.title

    def sheets(self):
        for ws in self.__wb:
            yield ws.title

    def rows(self):
        for row in self.__ws.iter_rows():
            yield [x.value for x in row]

class TestCaseConverter(object):
    def __init__(self, filename, output_dir):
        self.__reader = OpenpyxlReader(filename)
        self.__output_dir = output_dir
        self.__header = {}
        self.__clear()
        
    def load(self):
        self.__reader.load()

    def dump(self):
        self.__makedir()
        
        for sheetname in self.__reader.sheets():
            self.__dump_sheet(sheetname)
        
    def __clear(self):
        self.__givens = []
        self.__whens  = []
        self.__thens  = []

    def __makedir(self):
        if os.path.exists(self.__output_dir):
            shutil.rmtree(self.__output_dir)
        os.makedirs(self.__output_dir)

    def __dump_sheet(self, sheetname):
        self.__reader.activate_sheet(sheetname)

        if VERBOSE: print(sheetname)

        """
        0: ID
        1: Scenario
        2: Description
        3: Given - Part No
        4: Given - Belonged CB
        5: Given - Quantity
        6: Given - Web Price
        7: Given - Cost Price
        8: Given - Web Price Rule
        9: Given - Instant Saving Price Rule
        10: Given - Related CB
        11: Given - Floor Price Rule
        12: Given - Coupon Price Rule
        13: When - Action
        14: Then - Web Price
        15: Then - Sale Price
        16: Then - Floor Price
        17: Then - Coupon Price
        18: Then - Floor Price Violated by Sale Price
        19: Then - Floor Price Violated by Coupon Price
        """
        self.__parse_header()

        for i, items in enumerate(self.__reader.rows()):

            if self.__is_header(i):
                continue

            if VERBOSE: print(items)

            if self.__is_test_header(items[0]):
                if self.__has_test_body():
                    print("Generating test case {} ...".format(self.__feature_name))
                    self.__generate_gherkin()

                self.__feature_name         = self.__parse_feature_name(items[0], sheetname)
                self.__feature_description  = self.__parse_dsl_scalar(items[2])
                self.__test_case_name       = self.__parse_test_case_name(items[0], items[1])
                self.__clear()
            
            self.__givens.append(self.__parse_dsl_group(items, range(3, 13)))
            self.__whens .append(self.__parse_dsl_group(items, range(13, 14)))
            self.__thens .append("{}: {}".format(items[3], self.__parse_dsl_group(items, range(14, 20))))

        print("Generating test case {} ...".format(self.__feature_name))
        self.__generate_gherkin()
    
    def __parse_header(self):
        header1 = self.__reader[1]
        header2 = self.__reader[2]
        
        given_index = header1.index("Given")
        when_index  = header1.index("When")
        then_index  = header1.index("Then")

        for i in range(given_index, when_index)  : self.__header[i] = header2[i]
        for i in range(when_index , then_index)  : self.__header[i] = header2[i]
        for i in range(then_index , len(header1)): self.__header[i] = header2[i]

    def __is_header(self, index):
        return index < 2

    def __is_test_header(self, item):
        return False if item is None else re.search(r"\d+", str(item)) is not None
    
    def __has_test_body(self):
        return len(self.__givens) > 0

    def __parse_feature_name(self, id, name):
        return "{}_{}".format(name, id)

    def __parse_test_case_name(self, id, scenario):
        return "[{}] {}".format(id, "".join(self.__parse_dsl_scalar(scenario)))

    def __parse_dsl_scalar(self, item):
        return [x for x in item.split(FILE_NEWLINE) if x.strip() and x != NA] if item else []

    def __parse_dsl_group(self, items, indexes):
        return ", ".join(["{} is {}".format(self.__header[i], items[i]) for i in indexes if items[i]])
        
    def __generate_givens(self):
        return [GIVEN_TEMPLATE.format(x) for x in self.__givens if x.strip()]
        
    def __generate_whens(self):
        return [WHEN_TEMPLATE.format(x) for x in self.__whens if x.strip()]
    
    def __generate_thens(self):
        return [THEN_TEMPLATE.format(x) for x in self.__thens if x.strip()]
        
    def __generate_gherkin(self):
        sep_feature  = "{}{}".format(FILE_NEWLINE, FEATURE_INDENT)
        sep_scenario = "{}{}".format(FILE_NEWLINE, FEATURE_INDENT * 2)
        
        gherkin = FEATURE_TEMPLATE.format(
            self.__feature_name,
            sep_feature .join(self.__feature_description),
            self.__test_case_name,
            sep_scenario.join(self.__generate_givens()),
            sep_scenario.join(self.__generate_whens()),
            sep_scenario.join(self.__generate_thens()))
            
        if VERBOSE: print(gherkin)
        
        feature_filename = os.path.join(self.__output_dir, FEATURE_FILENAME.format(self.__feature_name))
        with codecs.open(feature_filename, "w", "utf-8") as feature_file:
            feature_file.write(gherkin)


def main():
    if len(sys.argv) != 2:
        print("[USAGE] python {} <filename>".format(os.path.basename(__file__)))
        sys.exit(1)
    
    filename   = sys.argv[1]
    output_dir = os.path.join(os.getcwd(), "features")
    
    converter = TestCaseConverter(filename, output_dir)
    converter.load()
    converter.dump()


def test_convert_test_case_from_excel_to_gherkin():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename   = os.path.join(script_dir, "testdata", "ecomm-cb-sbe-v0.5.xlsx")
    output_dir = os.path.join(os.getcwd(), "features")
    
    converter = TestCaseConverter(filename, output_dir)
    converter.load()
    converter.dump()
    
def test_read_excel_data():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename   = os.path.join(script_dir, "testdata", "ecomm-cb-sbe-v0.5.xlsx")

    reader = OpenpyxlReader(filename)
    reader.load()

    for name in reader.sheets(): print(name)
    
    print("Active sheet name: {}".format(reader.get_active_sheet()))
    for row in reader.rows(): print(row)

def test_excel_reader():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    testdata_dir = os.path.join(script_dir, "testdata")
    filename = os.path.join(testdata_dir, "helloworld.xlsx")

    if not os.path.exists(testdata_dir):
        os.makedirs(testdata_dir)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "Hello"
    ws["C2"] = "World"
    wb.create_sheet("Sheet2")
    wb.save(filename)

    reader = OpenpyxlReader(filename)
    reader.load()
    print("{}, {}".format(reader["B2"], reader["C2"]))
    print(reader[2])

    for name in reader.sheets(): print(name)
    
    print("Active sheet name: {}".format(reader.get_active_sheet()))
    for row in reader.rows(): print(row)

    reader.activate_sheet("Sheet2")

    print("Active sheet name: {}".format(reader.get_active_sheet()))
    for row in reader.rows(): print(row)
    
if __name__ == "__main__":
    main()
    #test_excel_reader()
    #test_read_excel_data()
    #test_convert_test_case_from_excel_to_gherkin()